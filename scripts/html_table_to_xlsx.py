#!/usr/bin/env python3
from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

from lxml import html as lxml_html
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FILL_COLORS = {
    "title": "8FB4D9",
    "breaktitle": "6F9FC9",
    "head": "DBEAF7",
    "hierhead": "D8E3EC",
    "camp": "D8E3EC",
    "adsetA": "E6EEF6",
    "adsetB": "F1F6FA",
    "dimtitle": "B9D4EC",
    "setting": "EEF5FB",
    "summary": "FFF4BF",
    "good": "D9EAD3",
    "bad": "F4CCCC",
    "normal": "FFFFFF",
    "padcell": "FFFFFF",
}

PIXEL_WIDTHS = {
    "pad": 24,
    "c1": 110,
    "c2": 210,
    "c3": 120,
    "c4": 112,
    "c5": 112,
    "c6": 92,
    "c7": 120,
    "c8": 110,
    "c9": 120,
    "c10": 95,
    "c11": 120,
    "c12": 110,
    "c13": 430,
}

THIN = Side(style="thin", color="D0D7DE")
MEDIUM = Side(style="medium", color="000000")


def pixel_to_width(px: int) -> float:
    # Close enough for Google Sheets import.
    return round((px - 5) / 7, 2)


def fill_for(classes: set[str]) -> PatternFill:
    for name, color in FILL_COLORS.items():
        if name in classes:
            return PatternFill(fill_type="solid", fgColor=color)
    return PatternFill(fill_type="solid", fgColor="FFFFFF")


def font_for(classes: set[str]) -> Font:
    bold_classes = {
        "title",
        "breaktitle",
        "head",
        "hierhead",
        "camp",
        "dimtitle",
        "setting",
        "summary",
        "good",
        "bad",
    }
    size = 11 if {"title", "breaktitle"} & classes else 10
    return Font(name="Arial", bold=bool(bold_classes & classes), size=size, color="1F2933")


def alignment_for(classes: set[str], text: str) -> Alignment:
    horizontal = "left"
    if classes & {"title", "breaktitle", "head", "hierhead", "camp", "center", "setting", "good", "bad", "normal"}:
        horizontal = "center"
    if "num" in classes and len(text) <= 20:
        horizontal = "center"
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=True)


def border_for(classes: set[str]) -> Border:
    return Border(
        left=MEDIUM if "left" in classes else THIN,
        right=MEDIUM if "right" in classes else THIN,
        top=MEDIUM if "top" in classes else THIN,
        bottom=MEDIUM if "bottom" in classes else THIN,
    )


def style_range(ws, row: int, col: int, rowspan: int, colspan: int, *, fill, font, alignment, border) -> None:
    for r in range(row, row + rowspan):
        for c in range(col, col + colspan):
            cell = ws.cell(r, c)
            cell.fill = copy(fill)
            cell.font = copy(font)
            cell.alignment = copy(alignment)
            cell.border = copy(border)


def parse_table(html_path: Path):
    root = lxml_html.fromstring(html_path.read_text(encoding="utf-8"))
    table = root.xpath("//table")[0]
    rows = table.xpath(".//tr")
    cols = [col.get("class", "") for col in table.xpath(".//col")]
    return rows, cols


def build_xlsx(html_path: Path, out_path: Path) -> None:
    rows, cols = parse_table(html_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    for idx, cls in enumerate(cols, start=1):
        width = PIXEL_WIDTHS.get(cls, 110)
        ws.column_dimensions[get_column_letter(idx)].width = pixel_to_width(width)

    occupied: set[tuple[int, int]] = set()
    max_col = 0

    for row_idx, tr in enumerate(rows, start=1):
        col_idx = 1
        for td in tr.xpath("./td"):
            while (row_idx, col_idx) in occupied:
                col_idx += 1

            rowspan = int(td.get("rowspan", "1"))
            colspan = int(td.get("colspan", "1"))
            classes = set((td.get("class") or "").split())
            text = " ".join(part.strip() for part in td.text_content().splitlines())
            text = " ".join(text.split())

            fill = fill_for(classes)
            font = font_for(classes)
            alignment = alignment_for(classes, text)
            border = border_for(classes)

            style_range(
                ws,
                row_idx,
                col_idx,
                rowspan,
                colspan,
                fill=fill,
                font=font,
                alignment=alignment,
                border=border,
            )
            ws.cell(row_idx, col_idx).value = text

            for r in range(row_idx, row_idx + rowspan):
                for c in range(col_idx, col_idx + colspan):
                    occupied.add((r, c))

            if rowspan > 1 or colspan > 1:
                ws.merge_cells(
                    start_row=row_idx,
                    start_column=col_idx,
                    end_row=row_idx + rowspan - 1,
                    end_column=col_idx + colspan - 1,
                )

            max_col = max(max_col, col_idx + colspan - 1)
            col_idx += colspan

    for row in range(1, len(rows) + 1):
        ws.row_dimensions[row].height = 20

    ws.freeze_panes = "A1"
    ws.sheet_view.showGridLines = True

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--html", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    build_xlsx(Path(args.html), Path(args.out))


if __name__ == "__main__":
    main()
